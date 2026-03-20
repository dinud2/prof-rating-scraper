import pandas as pd
import openpyxl
from openpyxl.styles import Font


def export_to_excel(matched, unmatched, output_path):
    column_map = {
        "code": "Course Code",
        "title": "Course Title",
        "year": "Year",
        "professor": "Professor Name",
        "rating": "Professor Rating",
        "difficulty": "Difficulty",
        "num_ratings": "Review Count",
    }

    matched_df = pd.DataFrame(matched).rename(columns=column_map)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        matched_df.to_excel(writer, sheet_name="Courses", index=False)
        if unmatched:
            unmatched_df = pd.DataFrame(unmatched).rename(columns=column_map)
            unmatched_df.to_excel(writer, sheet_name="Unmatched", index=False)

    wb = openpyxl.load_workbook(output_path)
    for ws in wb.worksheets:
        # Bold header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        # Freeze top row
        ws.freeze_panes = "A2"

        # Number format for rating columns (E = Professor Rating, F = Difficulty)
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "0.0"

    wb.save(output_path)
